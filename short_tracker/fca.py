"""FCA short-position downloads, immutable imports, and read APIs.

This module keeps the two disclosure regimes deliberately separate:

* the archived, named public disclosures under the pre-13 July 2026 regime;
* the FCA's anonymous aggregate net short positions (ANSP), first published on
  13 July 2026 for positions in scope from midnight on 9 July 2026.

The legacy issuer aggregate is reconstructed by replaying the exact disclosed
state of each ``(position holder, ISIN)``.  A published value of 0.50% or more
contributes; a later value below 0.50% removes that holder's contribution.
Source row order is retained and the last source row wins when the FCA workbook
contains more than one value for the same holder/issuer/ISIN/date.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from email.utils import parsedate_to_datetime
import csv
import hashlib
import json
import os
from pathlib import Path
import re
import sqlite3
import tempfile
import unicodedata
from typing import Any, Callable, Iterable, Iterator, Mapping
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from .db import Database


LEGACY_PUBLIC_THRESHOLD_BP = 50  # 0.50%, with one bp = 0.01 percentage point.
# The first ANSP publication on 13 July represented positions at midnight on
# 9 July. Keep the publication marker separate from the first position scope:
# the chart uses event/effective dates, while the UI's methodology marker stays
# on the day the new public output first appeared.
ANSP_INITIAL_POSITION_DATE = "2026-07-09"
ANSP_REGIME_START = "2026-07-13"
ANSP_FIRST_PUBLISHED_ON = ANSP_REGIME_START
IMPORTER_VERSION = 1
MAX_CURRENT_RANKING_LIMIT = 2_000


def _ansp_scope_start(rsl_date_added: str | None) -> str:
    """Return the first date on which an ISIN can be in ANSP chart scope.

    FCA's initial RSL was published/dated 13 July, but its first ANSP values
    represented positions at midnight on 9 July. Treat that initial cohort as
    in scope on 9 July. Shares added after the first publication enter scope on
    their actual RSL ``date_added``.
    """

    if not rsl_date_added or rsl_date_added <= ANSP_FIRST_PUBLISHED_ON:
        return ANSP_INITIAL_POSITION_DATE
    return max(ANSP_INITIAL_POSITION_DATE, rsl_date_added)


def _ansp_chart_schedule(
    historic_rows: Iterable[Mapping[str, Any]],
    *,
    current_position_date: str | None,
    rsl_date_added: str | None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """Link FCA ANSP states into step-after chart intervals.

    ``position_date`` is the latest constituent notification date and can move
    backwards after late or corrected notifications. It therefore determines
    only the initial state's lower bound. Historic states are ordered by the
    date each became historical; every later state begins exactly when the
    previous state ended. A current state begins at the last historic end, or
    at the initial scope/position bound when no history exists.
    """

    def source_row_number(row: Mapping[str, Any]) -> int:
        try:
            value = row["row_number"]
        except (KeyError, IndexError):
            return 0
        return int(value or 0)

    ordered = sorted(
        list(historic_rows),
        key=lambda row: (
            str(row["became_historical_date"]),
            source_row_number(row),
        ),
    )
    if not ordered and current_position_date is None:
        return [], None

    first_position_date = str(
        ordered[0]["position_date"] if ordered else current_position_date
    )
    scope_start = _ansp_scope_start(rsl_date_added)
    initial_chart_date = max(scope_start, first_position_date)
    historic_schedule: list[dict[str, Any]] = []
    previous_end: str | None = None
    for index, row in enumerate(ordered):
        position_date = str(row["position_date"])
        became_historical_date = str(row["became_historical_date"])
        chart_date = initial_chart_date if index == 0 else str(previous_end)
        if became_historical_date < chart_date:
            raise ValueError(
                "ANSP historical interval ends before it starts: "
                f"{chart_date} > {became_historical_date}"
            )
        historic_schedule.append(
            {
                "date": chart_date,
                "chart_date_basis": (
                    "initial_ansp_scope_and_constituent_position_date"
                    if index == 0
                    else "previous_became_historical_date"
                ),
                "chart_interpolation": "step_after",
                "ansp_scope_start": scope_start,
                "first_published_on": ANSP_FIRST_PUBLISHED_ON,
                "interval_end": became_historical_date,
                # Backwards-compatible diagnostic; chart_date_basis is the
                # authoritative explanation when date differs from position_date.
                "transition_date_clamped": chart_date != position_date,
            }
        )
        previous_end = became_historical_date

    current_schedule: dict[str, Any] | None = None
    if current_position_date is not None:
        chart_date = str(previous_end) if previous_end is not None else initial_chart_date
        current_schedule = {
            "date": chart_date,
            "chart_date_basis": (
                "previous_became_historical_date"
                if previous_end is not None
                else "initial_ansp_scope_and_constituent_position_date"
            ),
            "chart_interpolation": "step_after",
            "ansp_scope_start": scope_start,
            "first_published_on": ANSP_FIRST_PUBLISHED_ON,
            "interval_end": None,
            "transition_date_clamped": chart_date != current_position_date,
        }
    return historic_schedule, current_schedule


@dataclass(frozen=True)
class FCASource:
    key: str
    url: str
    file_name: str
    dataset_key: str | None = None


FCA_SOURCES: tuple[FCASource, ...] = (
    FCASource(
        "legacy_named_xlsx",
        "https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx",
        "short-positions-daily-update.xlsx",
        "legacy_named",
    ),
    FCASource(
        "ansp_current_csv",
        "https://www.fca.org.uk/publication/documents/aggregated-current-net-short-positions.csv",
        "aggregated-current-net-short-positions.csv",
        "ansp_current",
    ),
    FCASource(
        "ansp_historic_csv",
        "https://www.fca.org.uk/publication/documents/aggregated-historic-net-short-positions.csv",
        "aggregated-historic-net-short-positions.csv",
        "ansp_historic",
    ),
    FCASource(
        "ansp_combined_xlsx",
        "https://www.fca.org.uk/publication/documents/aggregated-net-short-positions.xlsx",
        "aggregated-net-short-positions.xlsx",
    ),
    FCASource(
        "rsl_csv",
        "https://www.fca.org.uk/publication/documents/uk-reportable-shares-list.csv",
        "uk-reportable-shares-list.csv",
        "reportable_shares",
    ),
    FCASource(
        "rsl_xlsx",
        "https://www.fca.org.uk/publication/documents/uk-reportable-shares-list.xlsx",
        "uk-reportable-shares-list.xlsx",
    ),
)

SOURCE_BY_KEY = {source.key: source for source in FCA_SOURCES}


class FCADataError(RuntimeError):
    """Base exception for malformed or unusable FCA data."""


class FCASyncError(FCADataError):
    """Raised when a sync cannot atomically activate all required datasets."""


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _iso_utc(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", "" if value is None else str(value)).casefold()
    text = re.sub(r"[^\w]+", " ", text, flags=re.UNICODE)
    return " ".join(text.split())


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def _clean_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return " ".join(text.replace("\u00a0", " ").split())


def _parse_date(value: Any, *, allow_blank: bool = False) -> str | None:
    if value is None or _clean_text(value) == "":
        if allow_blank:
            return None
        raise FCADataError("required date is blank")
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _clean_text(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError as exc:
        raise FCADataError(f"unrecognised date: {text!r}") from exc


def _position_to_bp(value: Any) -> int:
    """Convert a published percentage to exact hundredths of a percentage point."""

    text = _clean_text(value).removesuffix("%").strip()
    try:
        amount = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise FCADataError(f"invalid percentage: {value!r}") from exc
    bp = int((amount * 100).to_integral_value(rounding=ROUND_HALF_UP))
    if bp < 0:
        raise FCADataError(f"negative public short position: {value!r}")
    return bp


def _bp_to_pct(bp: int | None) -> float | None:
    return None if bp is None else float(Decimal(bp) / Decimal(100))


def _row_hash(values: Iterable[Any]) -> str:
    serialised = json.dumps(
        ["" if value is None else str(value) for value in values],
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(serialised).hexdigest()


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


class _IssuerResolver:
    """Resolve exact ISIN first, then an exact normalised issuer-name alias.

    No fuzzy company matching is attempted.  Name fallback is needed to group
    multiple share classes in the RSL under one issuer; ambiguous exact names
    are left as separate issuers.
    """

    def __init__(self, conn: sqlite3.Connection, now: str):
        self.conn = conn
        self.now = now
        self.by_isin: dict[str, int] = {}
        self.by_name: dict[str, int | None] = {}

    def resolve(self, name: str, isin: str, source_key: str, priority: int) -> int:
        name = _clean_text(name)
        isin = _clean_text(isin).upper()
        if not name:
            raise FCADataError("issuer name is blank")
        if not isin:
            raise FCADataError(f"ISIN is blank for {name}")
        normalised = _normalize_name(name)

        issuer_id = self.by_isin.get(isin)
        if issuer_id is None:
            row = self.conn.execute(
                "SELECT issuer_id FROM issuer_identifiers WHERE isin = ?", (isin,)
            ).fetchone()
            issuer_id = int(row["issuer_id"]) if row else None

        if issuer_id is None:
            cached = self.by_name.get(normalised, "missing")
            if cached != "missing":
                issuer_id = cached
            else:
                rows = self.conn.execute(
                    "SELECT DISTINCT issuer_id FROM issuer_aliases WHERE normalized_name = ? LIMIT 2",
                    (normalised,),
                ).fetchall()
                issuer_id = int(rows[0]["issuer_id"]) if len(rows) == 1 else None
                self.by_name[normalised] = issuer_id

        if issuer_id is None:
            cursor = self.conn.execute(
                """
                INSERT INTO issuers(
                    canonical_name, normalized_name, canonical_priority,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (name, normalised, priority, self.now, self.now),
            )
            issuer_id = int(cursor.lastrowid)
            self.by_name[normalised] = issuer_id

        self.conn.execute(
            """
            INSERT INTO issuer_identifiers(isin, issuer_id, first_seen_at, last_seen_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(isin) DO UPDATE SET last_seen_at = excluded.last_seen_at
            """,
            (isin, issuer_id, self.now, self.now),
        )
        self.conn.execute(
            """
            INSERT INTO issuer_aliases(
                issuer_id, name, normalized_name, source_key,
                first_seen_at, last_seen_at
            ) VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(issuer_id, name, source_key)
            DO UPDATE SET last_seen_at = excluded.last_seen_at
            """,
            (issuer_id, name, normalised, source_key, self.now, self.now),
        )
        self.conn.execute(
            """
            UPDATE issuers
               SET canonical_name = CASE WHEN ? >= canonical_priority THEN ? ELSE canonical_name END,
                   normalized_name = CASE WHEN ? >= canonical_priority THEN ? ELSE normalized_name END,
                   canonical_priority = MAX(canonical_priority, ?),
                   updated_at = ?
             WHERE id = ?
            """,
            (priority, name, priority, normalised, priority, self.now, issuer_id),
        )
        self.by_isin[isin] = issuer_id
        current = self.by_name.get(normalised)
        if current is None:
            self.by_name[normalised] = issuer_id
        elif current != issuer_id:
            self.by_name[normalised] = None
        return issuer_id


class FCADataService:
    """Download, archive, import, and query official FCA short-position data.

    Public API:

    * ``sync(force=False)`` -> serialisable sync summary;
    * ``search(query, limit=20)`` -> issuer/security matches;
    * ``get_security(id_or_isin)`` -> one issuer record;
    * ``get_short_series(id_or_isin)`` -> separate legacy and ANSP series.
    * ``get_current_rankings(limit=2000)`` -> current ANSP rows ranked by value.
    """

    def __init__(
        self,
        db: Database,
        data_dir: str | Path,
        *,
        opener: Callable[..., Any] | None = None,
        now: Callable[[], datetime] | None = None,
        timeout: int = 90,
    ):
        self.db = db
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self._opener = opener or urlopen
        self._now = now or _utc_now
        self.timeout = timeout

    def sync(self, force: bool = False) -> dict[str, Any]:
        """Synchronise all FCA sources and atomically activate successful imports.

        Identical bytes reuse the existing immutable snapshot and import.  A
        changed file gets a new SHA-256-addressed archive and new imported rows.
        Dataset heads move only after every required import succeeds.
        """

        started_at = _iso_utc(self._now())
        with self.db.transaction() as conn:
            cursor = conn.execute(
                "INSERT INTO sync_runs(started_at, status, force) VALUES (?, 'running', ?)",
                (started_at, int(force)),
            )
            sync_id = int(cursor.lastrowid)

        source_results: dict[str, dict[str, Any]] = {}
        import_results: dict[str, dict[str, Any]] = {}
        try:
            for source in FCA_SOURCES:
                source_results[source.key] = self._fetch_source(source, force=force)

            # RSL first gives the current authoritative display name; the exact
            # ISIN remains the primary identity key for all imports.
            import_order = (
                ("reportable_shares", "rsl_csv"),
                ("ansp_current", "ansp_current_csv"),
                ("ansp_historic", "ansp_historic_csv"),
                ("legacy_named", "legacy_named_xlsx"),
            )
            for dataset_key, source_key in import_order:
                source_result = source_results[source_key]
                import_results[dataset_key] = self._ensure_import(
                    dataset_key,
                    int(source_result["snapshot_id"]),
                    self.data_dir / source_result["archive_path"],
                )

            activated_at = _iso_utc(self._now())
            with self.db.transaction(immediate=True) as conn:
                for dataset_key, source_key in import_order:
                    source_result = source_results[source_key]
                    imported = import_results[dataset_key]
                    conn.execute(
                        """
                        INSERT INTO dataset_heads(
                            dataset_key, snapshot_id, import_run_id,
                            sync_run_id, activated_at
                        ) VALUES (?, ?, ?, ?, ?)
                        ON CONFLICT(dataset_key) DO UPDATE SET
                            snapshot_id = excluded.snapshot_id,
                            import_run_id = excluded.import_run_id,
                            sync_run_id = excluded.sync_run_id,
                            activated_at = excluded.activated_at
                        """,
                        (
                            dataset_key,
                            int(source_result["snapshot_id"]),
                            int(imported["import_run_id"]),
                            sync_id,
                            activated_at,
                        ),
                    )

            completed_at = _iso_utc(self._now())
            result = {
                "sync_id": sync_id,
                "status": "success",
                "started_at": started_at,
                "completed_at": completed_at,
                "force": bool(force),
                "changed": any(item["status"] == "downloaded" for item in source_results.values()),
                "sources": source_results,
                "imports": import_results,
            }
            with self.db.transaction() as conn:
                conn.execute(
                    """
                    UPDATE sync_runs
                       SET completed_at = ?, status = 'success', details_json = ?
                     WHERE id = ?
                    """,
                    (completed_at, _json(result), sync_id),
                )
            return result
        except Exception as exc:
            completed_at = _iso_utc(self._now())
            details = {"sources": source_results, "imports": import_results}
            with self.db.transaction() as conn:
                conn.execute(
                    """
                    UPDATE sync_runs
                       SET completed_at = ?, status = 'failed', details_json = ?, error = ?
                     WHERE id = ?
                    """,
                    (completed_at, _json(details), f"{type(exc).__name__}: {exc}", sync_id),
                )
            if isinstance(exc, FCADataError):
                raise
            raise FCASyncError(str(exc)) from exc

    def _fetch_source(self, source: FCASource, *, force: bool) -> dict[str, Any]:
        checked_at = _iso_utc(self._now())
        previous = self.db.query_one(
            """
            SELECT * FROM raw_snapshots
             WHERE source_key = ?
             ORDER BY last_checked_at DESC, id DESC
             LIMIT 1
            """,
            (source.key,),
        )
        headers = {
            "Accept": "*/*",
            "User-Agent": "ShortTracker/1.0 (+official FCA public-data synchroniser)",
        }
        previous_path: Path | None = None
        if previous:
            previous_path = self.data_dir / previous["archive_path"]
        if previous and previous_path and previous_path.exists() and not force:
            if previous["etag"]:
                headers["If-None-Match"] = previous["etag"]
            if previous["http_last_modified"]:
                headers["If-Modified-Since"] = previous["http_last_modified"]
        if force:
            headers["Cache-Control"] = "no-cache"

        request = Request(source.url, headers=headers)
        try:
            response = self._opener(request, timeout=self.timeout)
            try:
                body = response.read()
                response_headers = response.headers
                status_code = int(getattr(response, "status", 200) or 200)
            finally:
                close = getattr(response, "close", None)
                if close:
                    close()
        except HTTPError as exc:
            if exc.code == 304 and previous and previous_path and previous_path.exists():
                with self.db.transaction() as conn:
                    conn.execute(
                        "UPDATE raw_snapshots SET last_checked_at = ? WHERE id = ?",
                        (checked_at, int(previous["id"])),
                    )
                return self._snapshot_result(previous, status="not_modified", checked_at=checked_at)
            raise FCASyncError(f"FCA download failed for {source.key}: HTTP {exc.code}") from exc
        except OSError as exc:
            raise FCASyncError(f"FCA download failed for {source.key}: {exc}") from exc

        if status_code != 200:
            raise FCASyncError(f"FCA download failed for {source.key}: HTTP {status_code}")
        self._validate_payload(source, body)
        digest = hashlib.sha256(body).hexdigest()
        relative_path = Path("raw") / source.key / digest / source.file_name
        archive_path = self.data_dir / relative_path
        if not archive_path.exists():
            archive_path.parent.mkdir(parents=True, exist_ok=True)
            fd, temporary_name = tempfile.mkstemp(
                prefix=f".{source.file_name}.", suffix=".tmp", dir=archive_path.parent
            )
            try:
                with os.fdopen(fd, "wb") as output:
                    output.write(body)
                    output.flush()
                    os.fsync(output.fileno())
                os.replace(temporary_name, archive_path)
            except Exception:
                try:
                    os.unlink(temporary_name)
                except FileNotFoundError:
                    pass
                raise

        get_header = getattr(response_headers, "get", lambda key, default=None: default)
        content_type = str(get_header("Content-Type", "") or "").split(";", 1)[0].strip()
        last_modified = str(get_header("Last-Modified", "") or "") or None
        etag = str(get_header("ETag", "") or "") or None
        effective_date = None
        if last_modified:
            try:
                effective_date = parsedate_to_datetime(last_modified).date().isoformat()
            except (TypeError, ValueError, OverflowError):
                pass
        metadata = {"http_status": status_code}

        with self.db.transaction() as conn:
            row = conn.execute(
                "SELECT * FROM raw_snapshots WHERE source_key = ? AND sha256 = ?",
                (source.key, digest),
            ).fetchone()
            if row:
                conn.execute(
                    """
                    UPDATE raw_snapshots
                       SET last_checked_at = ?, source_url = ?, file_name = ?,
                           byte_size = ?, content_type = ?, archive_path = ?,
                           http_last_modified = COALESCE(?, http_last_modified),
                           etag = COALESCE(?, etag),
                           effective_date = COALESCE(?, effective_date),
                           metadata_json = ?
                     WHERE id = ?
                    """,
                    (
                        checked_at,
                        source.url,
                        source.file_name,
                        len(body),
                        content_type,
                        relative_path.as_posix(),
                        last_modified,
                        etag,
                        effective_date,
                        _json(metadata),
                        int(row["id"]),
                    ),
                )
                snapshot_id = int(row["id"])
                fetch_status = "unchanged"
            else:
                cursor = conn.execute(
                    """
                    INSERT INTO raw_snapshots(
                        source_key, source_url, file_name, sha256, byte_size,
                        content_type, archive_path, first_retrieved_at,
                        last_checked_at, http_last_modified, etag,
                        effective_date, metadata_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        source.key,
                        source.url,
                        source.file_name,
                        digest,
                        len(body),
                        content_type,
                        relative_path.as_posix(),
                        checked_at,
                        checked_at,
                        last_modified,
                        etag,
                        effective_date,
                        _json(metadata),
                    ),
                )
                snapshot_id = int(cursor.lastrowid)
                fetch_status = "downloaded"

        snapshot = self.db.query_one("SELECT * FROM raw_snapshots WHERE id = ?", (snapshot_id,))
        assert snapshot is not None
        return self._snapshot_result(snapshot, status=fetch_status, checked_at=checked_at)

    @staticmethod
    def _validate_payload(source: FCASource, body: bytes) -> None:
        if not body:
            raise FCADataError(f"empty FCA response for {source.key}")
        prefix = body[:512].lstrip().lower()
        if prefix.startswith(b"<!doctype html") or prefix.startswith(b"<html"):
            raise FCADataError(f"HTML received instead of {source.file_name}")
        if source.file_name.endswith(".xlsx") and not body.startswith(b"PK"):
            raise FCADataError(f"invalid XLSX payload for {source.key}")
        if source.file_name.endswith(".csv"):
            first_line = body[:2048].decode("utf-8-sig", errors="replace").splitlines()[0]
            if "," not in first_line:
                raise FCADataError(f"invalid CSV header for {source.key}")

    @staticmethod
    def _snapshot_result(
        snapshot: sqlite3.Row | Mapping[str, Any], *, status: str, checked_at: str
    ) -> dict[str, Any]:
        return {
            "snapshot_id": int(snapshot["id"]),
            "source_key": snapshot["source_key"],
            "url": snapshot["source_url"],
            "file_name": snapshot["file_name"],
            "sha256": snapshot["sha256"],
            "byte_size": int(snapshot["byte_size"]),
            "archive_path": snapshot["archive_path"],
            "first_retrieved_at": snapshot["first_retrieved_at"],
            "last_checked_at": checked_at,
            "http_last_modified": snapshot["http_last_modified"],
            "effective_date": snapshot["effective_date"],
            "status": status,
        }

    def _ensure_import(
        self, dataset_key: str, snapshot_id: int, path: Path
    ) -> dict[str, Any]:
        existing = self.db.query_one(
            """
            SELECT * FROM import_runs
             WHERE dataset_key = ? AND snapshot_id = ? AND importer_version = ?
            """,
            (dataset_key, snapshot_id, IMPORTER_VERSION),
        )
        if existing and existing["status"] == "success":
            return {
                "import_run_id": int(existing["id"]),
                "dataset_key": dataset_key,
                "snapshot_id": snapshot_id,
                "row_count": int(existing["row_count"] or 0),
                "profile": json.loads(existing["profile_json"] or "{}"),
                "reused": True,
            }

        started_at = _iso_utc(self._now())
        with self.db.transaction() as conn:
            if existing:
                import_run_id = int(existing["id"])
                conn.execute(
                    """
                    UPDATE import_runs
                       SET started_at = ?, completed_at = NULL, status = 'running',
                           row_count = NULL, profile_json = '{}', error = NULL
                     WHERE id = ?
                    """,
                    (started_at, import_run_id),
                )
            else:
                cursor = conn.execute(
                    """
                    INSERT INTO import_runs(
                        dataset_key, snapshot_id, importer_version,
                        started_at, status
                    ) VALUES (?, ?, ?, ?, 'running')
                    """,
                    (dataset_key, snapshot_id, IMPORTER_VERSION, started_at),
                )
                import_run_id = int(cursor.lastrowid)

        try:
            with self.db.transaction(immediate=True) as conn:
                if dataset_key == "legacy_named":
                    conn.execute("DELETE FROM legacy_issuer_aggregate WHERE snapshot_id = ?", (snapshot_id,))
                    conn.execute("DELETE FROM legacy_events WHERE snapshot_id = ?", (snapshot_id,))
                    profile = self._import_legacy(conn, path, snapshot_id)
                elif dataset_key == "ansp_current":
                    conn.execute("DELETE FROM ansp_current WHERE snapshot_id = ?", (snapshot_id,))
                    profile = self._import_ansp_current(conn, path, snapshot_id)
                elif dataset_key == "ansp_historic":
                    conn.execute("DELETE FROM ansp_historic WHERE snapshot_id = ?", (snapshot_id,))
                    profile = self._import_ansp_historic(conn, path, snapshot_id)
                elif dataset_key == "reportable_shares":
                    conn.execute("DELETE FROM rsl_entries WHERE snapshot_id = ?", (snapshot_id,))
                    profile = self._import_rsl(conn, path, snapshot_id)
                else:
                    raise FCADataError(f"unknown dataset: {dataset_key}")
                completed_at = _iso_utc(self._now())
                conn.execute(
                    """
                    UPDATE import_runs
                       SET completed_at = ?, status = 'success', row_count = ?,
                           profile_json = ?, error = NULL
                     WHERE id = ?
                    """,
                    (completed_at, int(profile["row_count"]), _json(profile), import_run_id),
                )
        except Exception as exc:
            with self.db.transaction() as conn:
                conn.execute(
                    """
                    UPDATE import_runs
                       SET completed_at = ?, status = 'failed', error = ?
                     WHERE id = ?
                    """,
                    (_iso_utc(self._now()), f"{type(exc).__name__}: {exc}", import_run_id),
                )
            raise
        return {
            "import_run_id": import_run_id,
            "dataset_key": dataset_key,
            "snapshot_id": snapshot_id,
            "row_count": int(profile["row_count"]),
            "profile": profile,
            "reused": False,
        }

    def _import_legacy(
        self, conn: sqlite3.Connection, path: Path, snapshot_id: int
    ) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - bundled runtime includes it.
            raise FCADataError("openpyxl is required to import the FCA legacy workbook") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration as exc:
                raise FCADataError("legacy workbook has no header row") from exc
            expected = (
                "positionholder",
                "nameofshareissuer",
                "isin",
                "netshortposition",
                "positiondate",
            )
            actual = tuple(_normalize_header(item) for item in headers[:5])
            if actual != expected:
                raise FCADataError(f"unexpected legacy headers: {headers!r}")

            now = _iso_utc(self._now())
            resolver = _IssuerResolver(conn, now)
            row_count = 0
            earliest: str | None = None
            latest: str | None = None
            exact_duplicate_rows = 0
            duplicate_rows = 0
            duplicate_keys: set[tuple[str, str, str, str]] = set()
            conflicting_keys: set[tuple[str, str, str, str]] = set()
            values_by_key: dict[tuple[str, str, str, str], set[int]] = {}
            seen_hashes: set[str] = set()
            samples: list[dict[str, Any]] = []

            for row_number, values in enumerate(rows, start=2):
                values = tuple(values[:5])
                if all(value is None or _clean_text(value) == "" for value in values):
                    continue
                holder = _clean_text(values[0])
                company = _clean_text(values[1])
                isin = _clean_text(values[2]).upper()
                if not holder:
                    raise FCADataError(f"blank position holder at legacy row {row_number}")
                position_bp = _position_to_bp(values[3])
                position_date = _parse_date(values[4])
                assert position_date is not None
                issuer_id = resolver.resolve(company, isin, "legacy_named_xlsx", 10)
                digest = _row_hash((holder, company, isin, position_bp, position_date))
                if digest in seen_hashes:
                    exact_duplicate_rows += 1
                seen_hashes.add(digest)

                key = (holder, company, isin, position_date)
                previous_values = values_by_key.get(key)
                if previous_values is not None:
                    duplicate_rows += 1
                    duplicate_keys.add(key)
                    if position_bp not in previous_values:
                        conflicting_keys.add(key)
                        if len(samples) < 20:
                            samples.append(
                                {
                                    "position_holder": holder,
                                    "company_name": company,
                                    "isin": isin,
                                    "position_date": position_date,
                                    "values_pct": sorted(
                                        [_bp_to_pct(item) for item in previous_values | {position_bp}]
                                    ),
                                    "winning_row_number": row_number,
                                    "winning_value_pct": _bp_to_pct(position_bp),
                                }
                            )
                    previous_values.add(position_bp)
                else:
                    values_by_key[key] = {position_bp}

                conn.execute(
                    """
                    INSERT INTO legacy_events(
                        snapshot_id, row_number, issuer_id, position_holder,
                        issuer_name, isin, position_bp, position_date, row_hash
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        snapshot_id,
                        row_number,
                        issuer_id,
                        holder,
                        company,
                        isin,
                        position_bp,
                        position_date,
                        digest,
                    ),
                )
                row_count += 1
                earliest = min(earliest, position_date) if earliest else position_date
                latest = max(latest, position_date) if latest else position_date
        finally:
            workbook.close()

        aggregate_rows = self._derive_legacy_aggregate(conn, snapshot_id)
        return {
            "row_count": row_count,
            "position_date_min": earliest,
            "position_date_max": latest,
            "exact_duplicate_row_count": exact_duplicate_rows,
            "duplicate_composite_row_count": duplicate_rows,
            "duplicate_composite_key_count": len(duplicate_keys),
            "conflicting_duplicate_key_count": len(conflicting_keys),
            "duplicate_resolution": "last_source_row_wins_within_holder_issuer_isin_position_date",
            "conflict_samples": samples,
            "derived_aggregate_row_count": aggregate_rows,
            "legacy_public_threshold_pct": 0.5,
        }

    @staticmethod
    def _derive_legacy_aggregate(conn: sqlite3.Connection, snapshot_id: int) -> int:
        """Replay the named disclosure archive into issuer-level state.

        The SQL order is intentionally ``position_date, row_number``.  The FCA
        archive has a handful of duplicate same-day keys; processing the later
        source row last makes the result deterministic and preserves corrections
        or closure values exactly as they appear in the official workbook.
        """

        state: dict[tuple[int, str, str], int] = {}
        issuer_totals: dict[int, int] = {}
        issuer_holder_counts: dict[int, int] = {}
        current_date: str | None = None
        touched: dict[int, int] = {}
        inserted = 0

        def flush() -> None:
            nonlocal inserted
            if current_date is None:
                return
            for issuer_id, event_count in touched.items():
                conn.execute(
                    """
                    INSERT INTO legacy_issuer_aggregate(
                        snapshot_id, issuer_id, position_date, aggregate_bp,
                        active_holder_count, event_count
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        snapshot_id,
                        issuer_id,
                        current_date,
                        issuer_totals.get(issuer_id, 0),
                        issuer_holder_counts.get(issuer_id, 0),
                        event_count,
                    ),
                )
                inserted += 1

        cursor = conn.execute(
            """
            SELECT issuer_id, position_holder, isin, position_bp,
                   position_date, row_number
              FROM legacy_events
             WHERE snapshot_id = ?
             ORDER BY position_date ASC, row_number ASC
            """,
            (snapshot_id,),
        )
        for row in cursor:
            position_date = row["position_date"]
            if current_date is not None and position_date != current_date:
                flush()
                touched = {}
            current_date = position_date
            issuer_id = int(row["issuer_id"])
            key = (issuer_id, row["position_holder"], row["isin"])
            old = state.pop(key, None)
            if old is not None:
                issuer_totals[issuer_id] = issuer_totals.get(issuer_id, 0) - old
                issuer_holder_counts[issuer_id] = issuer_holder_counts.get(issuer_id, 0) - 1
            new = int(row["position_bp"])
            if new >= LEGACY_PUBLIC_THRESHOLD_BP:
                state[key] = new
                issuer_totals[issuer_id] = issuer_totals.get(issuer_id, 0) + new
                issuer_holder_counts[issuer_id] = issuer_holder_counts.get(issuer_id, 0) + 1
            touched[issuer_id] = touched.get(issuer_id, 0) + 1
        flush()
        return inserted

    @staticmethod
    def _csv_records(path: Path) -> Iterator[tuple[int, dict[str, str]]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise FCADataError(f"CSV has no headers: {path.name}")
            for row_number, raw in enumerate(reader, start=2):
                if not any(_clean_text(value) for value in raw.values()):
                    continue
                yield row_number, {_normalize_header(key): value for key, value in raw.items()}

    def _import_ansp_current(
        self, conn: sqlite3.Connection, path: Path, snapshot_id: int
    ) -> dict[str, Any]:
        now = _iso_utc(self._now())
        resolver = _IssuerResolver(conn, now)
        count = 0
        earliest: str | None = None
        latest: str | None = None
        seen: set[tuple[str, str]] = set()
        duplicate_keys = 0
        for row_number, row in self._csv_records(path):
            company = _clean_text(row.get("nameofcompany"))
            isin = _clean_text(row.get("internationalsecuritiesidentificationnumberisin")).upper()
            bp = _position_to_bp(row.get("aggregatednetshortposition"))
            position_date = _parse_date(row.get("positiondate"))
            assert position_date is not None
            key = (isin, position_date)
            duplicate_keys += int(key in seen)
            seen.add(key)
            issuer_id = resolver.resolve(company, isin, "ansp_current_csv", 20)
            digest = _row_hash((company, isin, bp, position_date))
            conn.execute(
                """
                INSERT INTO ansp_current(
                    snapshot_id, row_number, issuer_id, company_name, isin,
                    aggregate_bp, position_date, row_hash
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (snapshot_id, row_number, issuer_id, company, isin, bp, position_date, digest),
            )
            count += 1
            earliest = min(earliest, position_date) if earliest else position_date
            latest = max(latest, position_date) if latest else position_date
        return {
            "row_count": count,
            "position_date_min": earliest,
            "position_date_max": latest,
            "duplicate_key_row_count": duplicate_keys,
        }

    def _import_ansp_historic(
        self, conn: sqlite3.Connection, path: Path, snapshot_id: int
    ) -> dict[str, Any]:
        now = _iso_utc(self._now())
        resolver = _IssuerResolver(conn, now)
        count = 0
        earliest: str | None = None
        latest: str | None = None
        became_min: str | None = None
        became_max: str | None = None
        seen: set[tuple[str, str, str, int]] = set()
        duplicate_keys = 0
        for row_number, row in self._csv_records(path):
            company = _clean_text(row.get("nameofcompany"))
            isin = _clean_text(row.get("internationalsecuritiesidentificationnumberisin")).upper()
            bp = _position_to_bp(row.get("aggregatednetshortposition"))
            position_date = _parse_date(row.get("positiondate"))
            became_date = _parse_date(row.get("datetheaggregatednetshortpositionbecamehistorical"))
            assert position_date is not None and became_date is not None
            key = (isin, position_date, became_date, bp)
            duplicate_keys += int(key in seen)
            seen.add(key)
            issuer_id = resolver.resolve(company, isin, "ansp_historic_csv", 20)
            digest = _row_hash((company, isin, bp, position_date, became_date))
            conn.execute(
                """
                INSERT INTO ansp_historic(
                    snapshot_id, row_number, issuer_id, company_name, isin,
                    aggregate_bp, position_date, became_historical_date, row_hash
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snapshot_id,
                    row_number,
                    issuer_id,
                    company,
                    isin,
                    bp,
                    position_date,
                    became_date,
                    digest,
                ),
            )
            count += 1
            earliest = min(earliest, position_date) if earliest else position_date
            latest = max(latest, position_date) if latest else position_date
            became_min = min(became_min, became_date) if became_min else became_date
            became_max = max(became_max, became_date) if became_max else became_date
        return {
            "row_count": count,
            "position_date_min": earliest,
            "position_date_max": latest,
            "became_historical_date_min": became_min,
            "became_historical_date_max": became_max,
            "duplicate_key_row_count": duplicate_keys,
        }

    def _import_rsl(
        self, conn: sqlite3.Connection, path: Path, snapshot_id: int
    ) -> dict[str, Any]:
        now = _iso_utc(self._now())
        resolver = _IssuerResolver(conn, now)
        count = 0
        earliest: str | None = None
        latest: str | None = None
        seen_isins: set[str] = set()
        duplicate_isins = 0
        main_count = 0
        for row_number, row in self._csv_records(path):
            isin = _clean_text(row.get("shareisin")).upper()
            company = _clean_text(row.get("companyname"))
            date_added = _parse_date(row.get("dateadded"), allow_blank=True)
            share_class = _clean_text(row.get("classofsharemainorotherclassofshares"))
            duplicate_isins += int(isin in seen_isins)
            seen_isins.add(isin)
            main_count += int(share_class.casefold().startswith("main"))
            issuer_id = resolver.resolve(company, isin, "rsl_csv", 30)
            digest = _row_hash((isin, company, date_added, share_class))
            conn.execute(
                """
                INSERT INTO rsl_entries(
                    snapshot_id, row_number, issuer_id, company_name, isin,
                    date_added, share_class, row_hash
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    snapshot_id,
                    row_number,
                    issuer_id,
                    company,
                    isin,
                    date_added,
                    share_class,
                    digest,
                ),
            )
            count += 1
            if date_added:
                earliest = min(earliest, date_added) if earliest else date_added
                latest = max(latest, date_added) if latest else date_added
        return {
            "row_count": count,
            "date_added_min": earliest,
            "date_added_max": latest,
            "main_share_count": main_count,
            "duplicate_isin_row_count": duplicate_isins,
        }

    def search(self, query: str, limit: int = 20) -> list[dict[str, Any]]:
        """Search issuer aliases and exact/prefix ISINs using indexed fields."""

        text = _clean_text(query)
        if not text or limit <= 0:
            return []
        limit = min(int(limit), 100)
        normalised = _normalize_name(text)
        isin_query = re.sub(r"\s+", "", text).upper()
        name_prefix = self._like_escape(normalised) + "%"
        name_contains = "%" + self._like_escape(normalised) + "%"
        isin_prefix = self._like_escape(isin_query) + "%"

        sql = """
            SELECT issuer_id, MIN(rank) AS rank
              FROM (
                    SELECT id AS issuer_id,
                           CASE
                             WHEN normalized_name = ? THEN 2
                             WHEN normalized_name LIKE ? ESCAPE '\\' THEN 3
                             ELSE 6
                           END AS rank
                      FROM issuers
                     WHERE normalized_name LIKE ? ESCAPE '\\'
                    UNION ALL
                    SELECT issuer_id,
                           CASE
                             WHEN normalized_name = ? THEN 2
                             WHEN normalized_name LIKE ? ESCAPE '\\' THEN 4
                             ELSE 6
                           END AS rank
                      FROM issuer_aliases
                     WHERE normalized_name LIKE ? ESCAPE '\\'
                    UNION ALL
                    SELECT issuer_id,
                           CASE WHEN isin = ? THEN 0 ELSE 1 END AS rank
                      FROM issuer_identifiers
                     WHERE isin LIKE ? ESCAPE '\\'
              ) candidates
             GROUP BY issuer_id
             ORDER BY rank ASC, issuer_id ASC
             LIMIT ?
        """
        rows = self.db.query_all(
            sql,
            (
                normalised,
                name_prefix,
                name_contains,
                normalised,
                name_prefix,
                name_contains,
                isin_query,
                isin_prefix,
                limit,
            ),
        )
        results: list[dict[str, Any]] = []
        for row in rows:
            security = self.get_security(int(row["issuer_id"]))
            if security:
                results.append(
                    {
                        "id": security["id"],
                        "name": security["name"],
                        "isin": security["primary_isin"],
                        "isins": security["isins"],
                        "reportable": security["reportable"],
                        "current_ansp": security["current_ansp"],
                        "rank": int(row["rank"]),
                    }
                )
        return results

    @staticmethod
    def _like_escape(value: str) -> str:
        return value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")

    @staticmethod
    def _resolve_issuer_id(conn: sqlite3.Connection, identifier: int | str) -> int | None:
        if isinstance(identifier, int) or (isinstance(identifier, str) and identifier.isdigit()):
            issuer_id = int(identifier)
            row = conn.execute("SELECT id FROM issuers WHERE id = ?", (issuer_id,)).fetchone()
            return issuer_id if row else None
        isin = _clean_text(identifier).upper()
        row = conn.execute(
            "SELECT issuer_id FROM issuer_identifiers WHERE isin = ?", (isin,)
        ).fetchone()
        return int(row["issuer_id"]) if row else None

    def get_security(self, identifier: int | str) -> dict[str, Any] | None:
        """Return issuer identity, aliases, RSL membership, and current ANSP."""

        with self.db.connection(readonly=True) as conn:
            issuer_id = self._resolve_issuer_id(conn, identifier)
            if issuer_id is None:
                return None
            issuer = conn.execute("SELECT * FROM issuers WHERE id = ?", (issuer_id,)).fetchone()
            if issuer is None:
                return None
            isins = [
                row["isin"]
                for row in conn.execute(
                    "SELECT isin FROM issuer_identifiers WHERE issuer_id = ? ORDER BY isin",
                    (issuer_id,),
                )
            ]
            aliases = [
                row["name"]
                for row in conn.execute(
                    """
                    SELECT DISTINCT name FROM issuer_aliases
                     WHERE issuer_id = ? ORDER BY name COLLATE NOCASE
                    """,
                    (issuer_id,),
                )
            ]
            rsl = [
                dict(row)
                for row in conn.execute(
                    """
                    SELECT r.isin, r.company_name, r.date_added, r.share_class
                      FROM rsl_entries r
                      JOIN dataset_heads h
                        ON h.dataset_key = 'reportable_shares'
                       AND h.snapshot_id = r.snapshot_id
                     WHERE r.issuer_id = ?
                     ORDER BY CASE WHEN lower(r.share_class) LIKE 'main%' THEN 0 ELSE 1 END,
                              r.isin
                    """,
                    (issuer_id,),
                )
            ]
            current_row = conn.execute(
                """
                SELECT a.aggregate_bp, a.position_date, a.isin
                  FROM ansp_current a
                  JOIN dataset_heads h
                    ON h.dataset_key = 'ansp_current'
                   AND h.snapshot_id = a.snapshot_id
                 WHERE a.issuer_id = ?
                 ORDER BY a.position_date DESC, a.row_number DESC
                 LIMIT 1
                """,
                (issuer_id,),
            ).fetchone()
            current_ansp = None
            if current_row:
                current_ansp = {
                    "value": _bp_to_pct(int(current_row["aggregate_bp"])),
                    "unit": "percent_of_issued_share_capital",
                    "position_date": current_row["position_date"],
                    "isin": current_row["isin"],
                }
            primary_isin = None
            if rsl:
                primary_isin = rsl[0]["isin"]
            elif isins:
                primary_isin = isins[0]
            return {
                "id": issuer_id,
                "name": issuer["canonical_name"],
                "normalized_name": issuer["normalized_name"],
                "primary_isin": primary_isin,
                "isins": isins,
                "aliases": aliases,
                "reportable": bool(rsl),
                "reportable_shares": rsl,
                "current_ansp": current_ansp,
            }

    def get_short_series(self, identifier: int | str) -> dict[str, Any] | None:
        """Return separate legacy-derived and official ANSP histories.

        All ``value`` fields are percentage points of issued share capital, not
        fractions. The ANSP list preserves both FCA date fields, but uses the
        effective state interval for ``date``: first scope/constituent date,
        then the previous state's ``became_historical_date``. Publication on
        13 July remains separate metadata and the methodology marker.
        """

        security = self.get_security(identifier)
        if security is None:
            return None
        issuer_id = int(security["id"])
        legacy: list[dict[str, Any]] = []
        ansp: list[dict[str, Any]] = []
        coverage: dict[str, Any] = {}
        with self.db.connection(readonly=True) as conn:
            heads = {
                row["dataset_key"]: int(row["snapshot_id"])
                for row in conn.execute("SELECT dataset_key, snapshot_id FROM dataset_heads")
            }
            legacy_snapshot = heads.get("legacy_named")
            if legacy_snapshot is not None:
                for row in conn.execute(
                    """
                    SELECT position_date, aggregate_bp, active_holder_count, event_count
                      FROM legacy_issuer_aggregate
                     WHERE snapshot_id = ? AND issuer_id = ?
                     ORDER BY position_date ASC
                    """,
                    (legacy_snapshot, issuer_id),
                ):
                    legacy.append(
                        {
                            "date": row["position_date"],
                            "position_date": row["position_date"],
                            "value": _bp_to_pct(int(row["aggregate_bp"])),
                            "unit": "percent_of_issued_share_capital",
                            "active_disclosed_holders": int(row["active_holder_count"]),
                            "source_event_count": int(row["event_count"]),
                            "regime": "legacy_named_public_disclosures",
                        }
                    )

            historic_rows: list[dict[str, Any]] = []
            historic_snapshot = heads.get("ansp_historic")
            if historic_snapshot is not None:
                historic_rows = [
                    dict(row)
                    for row in conn.execute(
                        """
                        SELECT position_date, became_historical_date,
                               aggregate_bp, isin, row_number
                          FROM ansp_historic
                         WHERE snapshot_id = ? AND issuer_id = ?
                         ORDER BY isin ASC, became_historical_date ASC, row_number ASC
                        """,
                        (historic_snapshot, issuer_id),
                    )
                ]

            current_snapshot = heads.get("ansp_current")
            current_row: dict[str, Any] | None = None
            if current_snapshot is not None:
                row = conn.execute(
                    """
                    SELECT position_date, aggregate_bp, isin, row_number
                      FROM ansp_current
                     WHERE snapshot_id = ? AND issuer_id = ?
                     ORDER BY position_date DESC, row_number DESC
                     LIMIT 1
                    """,
                    (current_snapshot, issuer_id),
                ).fetchone()
                current_row = dict(row) if row is not None else None

            rsl_dates: dict[str, str] = {}
            for share in security.get("reportable_shares", []):
                isin = str(share.get("isin") or "")
                date_added = share.get("date_added")
                if not isin or not date_added:
                    continue
                value = str(date_added)
                rsl_dates[isin] = min(rsl_dates.get(isin, value), value)

            historic_by_isin: dict[str, list[dict[str, Any]]] = {}
            for row in historic_rows:
                historic_by_isin.setdefault(str(row["isin"]), []).append(row)
            series_isins = set(historic_by_isin)
            if current_row is not None:
                series_isins.add(str(current_row["isin"]))

            current: dict[str, Any] | None = None
            for isin in sorted(series_isins):
                rows = historic_by_isin.get(isin, [])
                this_current = current_row if current_row and current_row["isin"] == isin else None
                historic_schedule, current_schedule = _ansp_chart_schedule(
                    rows,
                    current_position_date=(
                        str(this_current["position_date"])
                        if this_current is not None
                        else None
                    ),
                    rsl_date_added=rsl_dates.get(isin),
                )
                for row, chart in zip(rows, historic_schedule, strict=True):
                    ansp.append(
                        {
                            **chart,
                            "position_date": row["position_date"],
                            "became_historical_date": row["became_historical_date"],
                            "value": _bp_to_pct(int(row["aggregate_bp"])),
                            "unit": "percent_of_issued_share_capital",
                            "isin": isin,
                            "is_current": False,
                            "regime": "anonymous_fca_ansp",
                        }
                    )
                if this_current is not None and current_schedule is not None:
                    current = {
                        **current_schedule,
                        "position_date": this_current["position_date"],
                        "became_historical_date": None,
                        "value": _bp_to_pct(int(this_current["aggregate_bp"])),
                        "unit": "percent_of_issued_share_capital",
                        "isin": isin,
                        "is_current": True,
                        "regime": "anonymous_fca_ansp",
                    }
                    ansp.append(dict(current))

            ansp.sort(
                key=lambda point: (
                    point["date"],
                    point["isin"],
                    bool(point["is_current"]),
                )
            )

            for dataset_key in (
                "legacy_named",
                "ansp_current",
                "ansp_historic",
                "reportable_shares",
            ):
                coverage[dataset_key] = self._coverage(conn, dataset_key)

        return {
            "security": security,
            "legacy": legacy,
            "ansp": ansp,
            "current": current,
            "coverage": coverage,
            "methodology": {
                "unit": "percentage points of issued share capital",
                "legacy": (
                    "Derived issuer-level sum of exact named position-holder/ISIN state. "
                    "Values at or above 0.50% contribute; a later below-threshold or "
                    "closure row removes the contribution. Same-key/same-date conflicts "
                    "use the later row in the FCA source workbook."
                ),
                "ansp": (
                    "Official FCA anonymous issuer aggregate shown as step-after state "
                    "intervals. The first state begins from the 9 July 2026 position "
                    "scope (or a later RSL/constituent date); later states begin when "
                    "the previous value became historical. Publication first occurred "
                    "on 13 July and remains a separate methodology marker."
                ),
                "identity": (
                    "Exact ISIN first; exact normalised issuer-name alias only as a fallback. "
                    "No fuzzy company merge is performed."
                ),
                "caveats": [
                    "Neither series is total market short interest; sub-threshold and exempt positions are absent.",
                    "FCA may revise ANSP history after late, corrected, or verified notifications.",
                    "ANSP position_date is constituent metadata and may move backwards; it is not each later state's chart date.",
                    "If no current ANSP row exists, the last historic interval ends at became_historical_date and the chart remains a gap; it is never filled with zero.",
                    "Legacy and ANSP are different measurements and must not be connected even when both refer to 9 July 2026.",
                ],
            },
        }

    def get_current_rankings(self, limit: int = 2000) -> dict[str, Any]:
        """Return the active FCA current-ANSP snapshot in stable rank order.

        The ranking grain is exactly one published current ANSP row per ISIN.
        Rows that resolve to the same issuer are deliberately not summed: they
        may represent distinct reportable shares or share classes.  Percentage
        values are percentage points of issued share capital for that ISIN.
        """

        if isinstance(limit, bool) or not isinstance(limit, int):
            raise TypeError("limit must be an integer")
        limit = min(max(limit, 1), MAX_CURRENT_RANKING_LIMIT)

        with self.db.connection(readonly=True) as conn:
            coverage = self._coverage(conn, "ansp_current")
            if coverage is None:
                return {
                    "items": [],
                    "count": 0,
                    "total": 0,
                    "total_count": 0,
                    "as_of_date": None,
                    "limit": limit,
                    "truncated": False,
                    "source": None,
                    "coverage": None,
                    "methodology": self._current_ranking_methodology(),
                }

            snapshot_id = int(coverage["snapshot_id"])
            total_row = conn.execute(
                """
                SELECT COUNT(*) AS count, MAX(position_date) AS as_of_date
                  FROM ansp_current
                 WHERE snapshot_id = ?
                """,
                (snapshot_id,),
            ).fetchone()
            total_count = int(total_row["count"] or 0) if total_row else 0
            as_of_date = total_row["as_of_date"] if total_row else None
            rows = conn.execute(
                """
                SELECT row_number, issuer_id, company_name, isin,
                       aggregate_bp, position_date
                  FROM ansp_current
                 WHERE snapshot_id = ?
                 ORDER BY aggregate_bp DESC,
                          company_name COLLATE NOCASE ASC,
                          company_name ASC,
                          isin ASC,
                          row_number ASC
                 LIMIT ?
                """,
                (snapshot_id, limit),
            ).fetchall()

        items = [
            {
                "rank": rank,
                "company": row["company_name"],
                "company_name": row["company_name"],
                "issuer_id": int(row["issuer_id"]),
                "security_id": int(row["issuer_id"]),
                "name": row["company_name"],
                "isin": row["isin"],
                "aggregate_bp": int(row["aggregate_bp"]),
                "percent": _bp_to_pct(int(row["aggregate_bp"])),
                "short_percent": _bp_to_pct(int(row["aggregate_bp"])),
                "position_date": row["position_date"],
                "source_row_number": int(row["row_number"]),
                "row_number": int(row["row_number"]),
                "unit": "percent_of_issued_share_capital",
            }
            for rank, row in enumerate(rows, start=1)
        ]
        return {
            "items": items,
            "count": len(items),
            "total": total_count,
            "total_count": total_count,
            "as_of_date": as_of_date,
            "limit": limit,
            "truncated": len(items) < total_count,
            "source": {
                "authority": "UK Financial Conduct Authority",
                "dataset": "aggregated_current_net_short_positions",
                "source_key": coverage["source_key"],
                "url": coverage["url"],
                "snapshot_id": snapshot_id,
                "sha256": coverage["sha256"],
            },
            "coverage": coverage,
            "methodology": self._current_ranking_methodology(),
        }

    @staticmethod
    def _current_ranking_methodology() -> dict[str, Any]:
        return {
            "ranking_unit": "one FCA current ANSP row per reportable share/ISIN",
            "aggregation": (
                "No aggregation is performed across ISINs or share classes, "
                "including rows that resolve to the same issuer."
            ),
            "ranking_order": (
                "aggregate_bp descending; ties by company name (case-insensitive, "
                "then exact text), ISIN, and FCA source row number"
            ),
            "position_date": (
                "Under FCA SSR 6.3.1, position_date is the most recent position "
                "date among the notifications included in that aggregate; an "
                "older date can therefore be valid in the current snapshot."
            ),
            "unit": "percentage points of issued share capital for the listed ISIN",
            "base_notification_threshold_percent": 0.2,
            "is_total_market_short_interest": False,
            "scope": (
                "The current ANSP report aggregates net short positions reported "
                "to and included by the FCA, normally individual positions at or "
                "above the 0.20% base notification threshold."
            ),
            "caveats": [
                "Sub-threshold, exempt, and otherwise undisclosed positions are absent.",
                "Omission from the current ANSP report does not prove zero short interest.",
                "FCA may amend or republish figures after late, corrected, or verified notifications.",
            ],
        }

    @staticmethod
    def _coverage(conn: sqlite3.Connection, dataset_key: str) -> dict[str, Any] | None:
        row = conn.execute(
            """
            SELECT h.dataset_key, h.activated_at, s.id AS snapshot_id,
                   s.source_key, s.source_url, s.file_name, s.sha256,
                   s.byte_size, s.first_retrieved_at, s.last_checked_at,
                   s.http_last_modified, s.effective_date,
                   i.row_count, i.profile_json, i.completed_at AS imported_at
              FROM dataset_heads h
              JOIN raw_snapshots s ON s.id = h.snapshot_id
              JOIN import_runs i ON i.id = h.import_run_id
             WHERE h.dataset_key = ?
            """,
            (dataset_key,),
        ).fetchone()
        if row is None:
            return None
        return {
            "dataset": dataset_key,
            "snapshot_id": int(row["snapshot_id"]),
            "source_key": row["source_key"],
            "url": row["source_url"],
            "file_name": row["file_name"],
            "sha256": row["sha256"],
            "byte_size": int(row["byte_size"]),
            "first_retrieved_at": row["first_retrieved_at"],
            "last_checked_at": row["last_checked_at"],
            "http_last_modified": row["http_last_modified"],
            "effective_date": row["effective_date"],
            "activated_at": row["activated_at"],
            "imported_at": row["imported_at"],
            "row_count": int(row["row_count"] or 0),
            "profile": json.loads(row["profile_json"] or "{}"),
        }
