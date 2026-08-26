"""Windowless entry point for the standalone Windows distribution."""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
import json
import os
from pathlib import Path
import sqlite3
import ssl
import sys
import tempfile
import traceback


SERVICE_CHILD_ARGUMENT = "--service-child"


def _service_arguments(arguments: list[str]) -> tuple[argparse.Namespace, list[str]]:
    parser = argparse.ArgumentParser(add_help=False, allow_abbrev=False)
    parser.add_argument("--stdout-log", type=Path)
    parser.add_argument("--stderr-log", type=Path)
    return parser.parse_known_args(arguments)


def _bind_service_streams(stdout_log: Path | None, stderr_log: Path | None) -> None:
    """Provide real text streams in a PyInstaller windowed child process."""

    if stdout_log is None or stderr_log is None:
        from short_tracker.paths import default_data_dir

        logs_dir = default_data_dir() / "runtime" / "logs"
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        fallback = logs_dir / f"{stamp}-{os.getpid()}-service-child"
        stdout_log = stdout_log or fallback.with_suffix(".out.log")
        stderr_log = stderr_log or fallback.with_suffix(".err.log")

    stdout_path = stdout_log.expanduser().resolve()
    stderr_path = stderr_log.expanduser().resolve()
    stdout_path.parent.mkdir(parents=True, exist_ok=True)
    stderr_path.parent.mkdir(parents=True, exist_ok=True)
    sys.stdout = stdout_path.open("a", encoding="utf-8", errors="replace", buffering=1)
    sys.stderr = stderr_path.open("a", encoding="utf-8", errors="replace", buffering=1)


def _show_fatal_error(error: BaseException) -> None:
    """Surface otherwise-silent failures from the windowed executable."""

    detail = "".join(traceback.format_exception(type(error), error, error.__traceback__))
    try:
        from short_tracker.paths import default_data_dir

        log_path = default_data_dir() / "runtime" / "logs" / "launcher-fatal.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8", errors="replace") as handle:
            handle.write(detail)
            handle.write("\n")
    except Exception:
        log_path = None

    message = (
        "Short Tracker 无法启动。\n\n"
        "Short Tracker could not start.\n\n"
        f"{type(error).__name__}: {error}"
        + (f"\n\nLog: {log_path}" if log_path is not None else "")
    )
    if os.name == "nt":
        try:
            import ctypes

            ctypes.windll.user32.MessageBoxW(0, message, "Short Tracker", 0x10)
            return
        except Exception:
            pass
    stream = sys.stderr or sys.__stderr__
    if stream is not None:
        stream.write(message + "\n")


def _bundle_self_test(arguments: list[str]) -> int:
    """Exercise dynamically bundled Windows dependencies without network use."""

    parser = argparse.ArgumentParser(add_help=False, allow_abbrev=False)
    parser.add_argument("--result-json", type=Path, required=True)
    args = parser.parse_args(arguments)
    payload: dict[str, object]
    try:
        from launcher.desktop_launcher import _configure_pythonnet_runtime

        _configure_pythonnet_runtime()
        import clr
        import webview
        from webview.guilib import initialize
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as temporary:
            workbook_path = Path(temporary) / "synthetic.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "Short Tracker"
            workbook.save(workbook_path)
            workbook.close()
            loaded = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                workbook_value = loaded.active["A1"].value
            finally:
                loaded.close()

        with sqlite3.connect(":memory:") as connection:
            sqlite_value = connection.execute("SELECT 1").fetchone()[0]
        ssl_context = ssl.create_default_context()
        webview_gui = initialize("edgechromium")
        renderer = str(webview_gui.renderer)
        payload = {
            "ok": (
                workbook_value == "Short Tracker"
                and sqlite_value == 1
                and renderer == "edgechromium"
            ),
            "openpyxl_workbook": workbook_value,
            "sqlite": sqlite_value,
            "ssl_protocol": int(ssl_context.protocol),
            "pythonnet": bool(clr),
            "pywebview": bool(webview),
            "webview_renderer": renderer,
        }
    except BaseException as error:
        payload = {
            "ok": False,
            "error": f"{type(error).__name__}: {error}",
            "traceback": "".join(
                traceback.format_exception(type(error), error, error.__traceback__)
            ),
        }
    args.result_json.parent.mkdir(parents=True, exist_ok=True)
    args.result_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return 0 if payload.get("ok") is True else 1


def _desktop_smoke_test(arguments: list[str]) -> int:
    """Run the real desktop shell on an isolated address for release QA."""

    parser = argparse.ArgumentParser(add_help=False, allow_abbrev=False)
    parser.add_argument("--data-dir", type=Path, required=True)
    parser.add_argument("--port", type=int, required=True)
    args = parser.parse_args(arguments)
    from launcher.desktop_launcher import ShortTrackerLauncher, run_desktop

    launcher = ShortTrackerLauncher(data_dir=args.data_dir, port=args.port)
    return run_desktop(launcher=launcher, skip_startup_sync=True)


def main(argv: list[str] | None = None) -> int:
    arguments = list(sys.argv[1:] if argv is None else argv)

    if arguments and arguments[0] == SERVICE_CHILD_ARGUMENT:
        internal, service_arguments = _service_arguments(arguments[1:])
        _bind_service_streams(internal.stdout_log, internal.stderr_log)
        try:
            from short_tracker.__main__ import main as service_main

            return service_main(service_arguments)
        except BaseException:
            traceback.print_exc(file=sys.stderr)
            return 1

    if arguments and arguments[0] == "--launcher-cli":
        internal, launcher_arguments = _service_arguments(arguments[1:])
        _bind_service_streams(internal.stdout_log, internal.stderr_log)
        from launcher.desktop_launcher import main as launcher_main

        return launcher_main(launcher_arguments)

    if arguments and arguments[0] == "--bundle-self-test":
        return _bundle_self_test(arguments[1:])

    if arguments and arguments[0] == "--desktop-smoke-test":
        return _desktop_smoke_test(arguments[1:])

    from launcher.desktop_launcher import run_desktop

    return run_desktop()


if __name__ == "__main__":
    try:
        exit_code = main()
    except BaseException as fatal_error:
        _show_fatal_error(fatal_error)
        exit_code = 1
    raise SystemExit(exit_code)
